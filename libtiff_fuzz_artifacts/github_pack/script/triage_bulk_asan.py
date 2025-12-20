#!/usr/bin/env python3
import argparse, os, re, shlex, subprocess, time
from pathlib import Path

def run_one(cmd, env, timeout_s):
    try:
        p = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            env=env,
            timeout=timeout_s,
            text=True,
            errors="replace",
        )
        return p.returncode, p.stdout, False
    except subprocess.TimeoutExpired as e:
        out = (e.stdout or "") + "\n[TIMEOUT]\n"
        return 124, out, True

def parse_sanitizer(output: str):
    # Error type
    m = re.search(r"ERROR:\s+AddressSanitizer:\s+([A-Za-z0-9_\-]+)", output)
    if m:
        err = "ASan: " + m.group(1)
    else:
        m = re.search(r"SUMMARY:\s+UndefinedBehaviorSanitizer:\s+([^\n]+)", output)
        if m:
            err = "UBSan: " + m.group(1).strip()
        else:
            m = re.search(r"SUMMARY:\s+AddressSanitizer:\s+([^\n]+)", output)
            err = ("ASan: " + m.group(1).strip()) if m else "Unknown"

    # Top frame: line beginning with #0
    func = ""
    src = ""
    m = re.search(r"^\s*#0\s+0x[0-9a-f]+\s+in\s+([^\s]+)\s+(.+?):(\d+):(\d+)", output, re.M)
    if m:
        func = m.group(1)
        src = f"{m.group(2)}:{m.group(3)}:{m.group(4)}"
    else:
        # fallback: "#0 ... in FUNC /path:line:col"
        m = re.search(r"^\s*#0\s+.*\sin\s+([^\s]+)\s+(.+)$", output, re.M)
        if m:
            func = m.group(1)
            src = m.group(2).strip()

    return err, func, src

def safe_name(p: Path):
    # turn "id:000123,..." into safe filename
    return re.sub(r"[^A-Za-z0-9._,-]+", "_", p.name)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--crash-dir", required=True)
    ap.add_argument("--pattern", default="id:*")
    ap.add_argument("--target", required=True)
    ap.add_argument("--mode", choices=["tiffcp", "tiff2pdf"], required=True)
    ap.add_argument("--timeout", type=int, default=5)
    ap.add_argument("--limit", type=int, default=0, help="0 = no limit")
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--keep-logs", action="store_true", help="keep per-file logs")
    args = ap.parse_args()

    crash_dir = Path(args.crash_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    logs_dir = out_dir / "logs"
    if args.keep_logs:
        logs_dir.mkdir(parents=True, exist_ok=True)

    # Env for ASan
    env = os.environ.copy()
    env["ASAN_OPTIONS"] = env.get("ASAN_OPTIONS", "") + ":abort_on_error=1:symbolize=1:detect_leaks=0"
    env["UBSAN_OPTIONS"] = env.get("UBSAN_OPTIONS", "") + ":print_stacktrace=1"

    # Collect inputs
    inputs = sorted(crash_dir.glob(args.pattern))
    inputs = [p for p in inputs if p.is_file()]
    if args.limit and args.limit > 0:
        inputs = inputs[:args.limit]

    rows = []
    for i, inp in enumerate(inputs, 1):
        if args.mode == "tiffcp":
            cmd = [args.target, str(inp), "/tmp/out.tif"]
            repro = f"{shlex.quote(args.target)} {shlex.quote(str(inp))} /tmp/out.tif"
        else:
            cmd = [args.target, "-o", "/tmp/out.pdf", str(inp)]
            repro = f"{shlex.quote(args.target)} -o /tmp/out.pdf {shlex.quote(str(inp))}"

        rc, out, timed = run_one(cmd, env, args.timeout)
        err, func, src = parse_sanitizer(out)

        log_path = ""
        if args.keep_logs:
            log_path = str(logs_dir / (safe_name(inp) + ".log"))
            Path(log_path).write_text(out)

        # Suggested search query string (for Google)
        # include func + file hint when available
        query_bits = []
        if func:
            query_bits.append(func)
        if src and ":" in src:
            query_bits.append(src.split(":")[0].split("/")[-1])
        query_bits.append("libtiff")
        query_bits.append("CVE")
        google_q = " ".join(query_bits)

        rows.append({
            "target": args.mode,
            "input_file": str(inp),
            "result": "timeout" if timed else ("crash" if ("AddressSanitizer" in out or "UndefinedBehaviorSanitizer" in out) else "ok"),
            "sanitizer_error": err,
            "top_function": func,
            "source_hint": src,
            "return_code": rc,
            "repro_cmd": repro,
            "log_path": log_path,
            "google_query": google_q,
            "suspected_cve": "",
        })

        if i % 50 == 0:
            print(f"[+] Processed {i}/{len(inputs)} ...")

    # Write xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    except Exception:
        raise SystemExit("Missing openpyxl. Install: python3 -m pip install --user openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"triage_{args.mode}"

    headers = [
        "target", "input_file", "result", "sanitizer_error", "top_function",
        "source_hint", "return_code", "repro_cmd", "log_path",
        "google_query", "suspected_cve"
    ]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in rows:
        ws.append([r[h] for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # widths
    widths = {
        "A": 10, "B": 55, "C": 10, "D": 22, "E": 28, "F": 40,
        "G": 10, "H": 60, "I": 45, "J": 35, "K": 18
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_xlsx = Path(args.xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"[+] Wrote: {out_xlsx}")

if __name__ == "__main__":
    main()
